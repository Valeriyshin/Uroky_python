# # Задание 1 — Калькулятор деления (исключения)
# # Сделать программу, которая:
# # 1) Просит у пользователя два числа a и b.
# # 2) Печатает результат a / b.
# # 3) Обрабатывает ошибки: ValueError (ввод не число), ZeroDivisionError (b=0).
# # 4) В finally печатает: 'Завершено'.
# # Требование: отдельные except для ValueError и ZeroDivisionError.
# print("Калькулятор деления")
# try:
#     a = int(input("Введите число a: "))
#     b = int(input("Введите число b: "))
#     result = a/b
#     print("Результат: ", result)
# except ValueError:
#     print("Ошибка: введено не число")
# except ZeroDivisionError:
#     print("Ошибка: нельзя делить на ноль")
# finally: 
#     print("завершено")

# # Задание 2 — Excel: товары и отчёт
# # Создай файл products.xlsx со столбцами: ID, Title, Price, Quantity.
# # Напиши скрипт, который:
# # 1) Читает products.xlsx.
# # 2) Для каждой строки делает объект Product.
# # 3) Если Price/Quantity пустые или кривые — строку пропускает (через except (ValueError, TypeError)).
# # 4) Сохраняет products_report.xlsx с колонками: Title, LineTotal.
# # 5) Внизу отчёта добавляет строку 'TOTAL' и общую сумму.

# from openpyxl import Workbook, load_workbook
# class Product:
#     def __init__(self, title, price, quantity):
#         self.title = title
#         self.price = price
#         self.quantity = quantity
#     def line_total(self):
#         return self.price * self.quantity
    
# wb = Workbook()
# sheet = wb.active
# sheet.title = "Products"

# sheet.append(["ID", "Title", "Price", "Quantity"])
# sheet.append([1, "Iphone", 1100, 2])
# sheet.append([2, "Samsung", 900, 3])
# sheet.append([3, "Xiaomi", "—", 5])
# sheet.append([4, "Nokia", 500, None])

# wb.save("products.xlsx")
# wb = load_workbook("products.xlsx")
# sheet = wb.active
# products = []

# for row in sheet.iter_rows(min_row=2, values_only=True):
#     id, title, price, quantity = row
#     try:
#         price = float(price)
#         quantity = int(quantity)
#         product = Product(title, price, quantity)
#         products.append(product)
#     except (ValueError, TypeError):
#         print(f"Пропущена строка с ID {id} из-за некорректных данных.")
#         continue

# report_wb = Workbook()
# report_sheet = report_wb.active
# report_sheet.title = "Products Report"

# report_sheet.append(["Title", "LineTotal"])
# total_sum = 0
# for product in products:
#     line_total = product.line_total()
#     report_sheet.append([product.title, line_total])
#     total_sum += line_total
# report_sheet.append(["TOTAL", total_sum])
# report_wb.save("products_report.xlsx")


# Задание 3 — Excel: студенты и средний балл
# Создай students2.xlsx со столбцами: ID, Name, Score1, Score2, Score3.
# Скрипт должен:
# 1) Для каждой строки создать объект Student.
# 2) Добавить оценки через add_score.
# 3) Посчитать avg() и сохранить report_students.xlsx: Name, Avg.
# 4) Обработать кривые значения (None, '—', 'abc') через (ValueError, TypeError).
# Чек-лист перед сдачей (чтобы студенты не забыли)
# ☐ Проект запускается без ошибок (python main.py).
# ☐ Есть обработка ошибок через try/except (не падает на грязных данных).
# ☐ Excel-файлы читаются и создаются корректно.
# ☐ Код разбит на функции/класс (не всё в одной простыне).
# ☐ Есть README: как установить openpyxl и как запустить.

from openpyxl import Workbook, load_workbook
class Student:
    def __init__(self, name):
        self.name = name
        self.scores = []
    def add_score(self, score):
        self.scores.append(score)
    def avg(self):
        return sum(self.scores) / len(self.scores) if self.scores else 0
wb = Workbook()
sheet = wb.active
sheet.title = "Students"
sheet.append(["ID", "Name", "Score1", "Score2", "Score3"])
sheet.append([1, "Alice", 85, 90, 78])
sheet.append([2, "Bob", 88, "—", 92])
sheet.append([3, "Charlie", None, 70, 75])
sheet.append([4, "David", 95, 85, "abc"])
wb.save("students2.xlsx")
wb = load_workbook("students2.xlsx")
sheet = wb.active
students = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    id, name, score1, score2, score3 = row
    student = Student(name)
    for score in (score1, score2, score3):
        try:
            score = float(score)
            student.add_score(score)
        except (ValueError, TypeError):
            print(f"Пропущена оценка для студента {name} из-за некорректных данных.")
            continue
    students.append(student)
report_wb = Workbook()
report_sheet = report_wb.active
report_sheet.title = "Students Report"
report_sheet.append(["Name", "Avg"])
for student in students:
    average = student.avg()
    report_sheet.append([student.name, average])
report_wb.save("report_students.xlsx")
